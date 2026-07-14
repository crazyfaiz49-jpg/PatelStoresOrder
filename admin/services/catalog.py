import csv
import json
import shutil
import subprocess
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from PIL import Image

from admin.database.db import DB_PATH, ROOT_DIR, get_connection
from admin.utils.logger import get_logger

logger = get_logger(__name__)
IMAGES_DIR = ROOT_DIR / 'images'
IMAGES_DIR.mkdir(exist_ok=True)
BACKUP_DIR = ROOT_DIR / 'backup'
BACKUP_DIR.mkdir(exist_ok=True)


@dataclass
class ImportReport:
    added: int = 0
    updated: int = 0
    skipped: int = 0
    duplicate: int = 0
    errors: int = 0
    total: int = 0
    elapsed_seconds: float = 0.0
    report_path: str = ''


class PublishError(RuntimeError):
    pass


def _safe_float(value, default: float = 0.0) -> float:
    try:
        if value in (None, ''):
            return float(default)
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def slugify(value: str) -> str:
    import re

    value = re.sub(r'[^a-z0-9]+', '-', (value or '').lower()).strip('-')
    return value or 'product'


def normalize_image_value(image_value: Optional[str]) -> str:
    if not image_value:
        return ''
    path = str(image_value).strip()
    if path.startswith('images/'):
        path = path.split('/', 1)[1]
    return Path(path).name


def get_catalog_image_path(image_value: Optional[str]) -> str:
    file_name = normalize_image_value(image_value)
    if not file_name:
        return 'images/placeholder.png'
    return f'images/{file_name}'


def build_image_filename(product_name: str, input_name: str = '') -> str:
    suffix = Path(input_name or '').suffix.lower()
    if suffix not in {'.jpg', '.jpeg', '.png'}:
        suffix = '.jpg'
    if suffix == '.jpeg':
        suffix = '.jpg'
    return f'{slugify(product_name)}{suffix}'


def save_image_from_path(source_path: str, product_name: str) -> str:
    src = Path(source_path)
    if not src.exists() or not src.is_file():
        return ''

    target_name = build_image_filename(product_name, src.name)
    target_path = IMAGES_DIR / target_name
    if target_path.exists():
        target_path = IMAGES_DIR / f"{slugify(product_name)}-{datetime.now().strftime('%H%M%S')}{target_path.suffix}"

    with Image.open(src) as image:
        if target_path.suffix.lower() in {'.jpg', '.jpeg'}:
            image = image.convert('RGB')
        image.thumbnail((800, 800))
        save_args = {'optimize': True}
        if target_path.suffix.lower() in {'.jpg', '.jpeg'}:
            save_args['quality'] = 82
        image.save(target_path, **save_args)

    return target_path.name


def optimize_image_copy(source_path: Path, target_path: Path):
    target_path.parent.mkdir(parents=True, exist_ok=True)
    suffix = target_path.suffix.lower()
    if suffix in {'.jpg', '.jpeg', '.png'}:
        with Image.open(source_path) as image:
            if suffix in {'.jpg', '.jpeg'}:
                image = image.convert('RGB')
            image.thumbnail((1600, 1600))
            save_args = {'optimize': True}
            if suffix in {'.jpg', '.jpeg'}:
                save_args['quality'] = 82
            image.save(target_path, **save_args)
        return

    shutil.copy2(source_path, target_path)


def save_sqlite_database():
    connection = get_connection()
    try:
        connection.commit()
        try:
            connection.execute('PRAGMA wal_checkpoint(FULL)')
        except Exception:
            logger.debug('SQLite wal_checkpoint skipped', exc_info=True)
    finally:
        connection.close()


def _git_run(args: List[str], *, allow_failure: bool = False) -> subprocess.CompletedProcess:
    proc = subprocess.run(['git', *args], cwd=ROOT_DIR, check=False, capture_output=True, text=True)
    if proc.returncode != 0 and not allow_failure:
        message = (proc.stderr or proc.stdout or f"git {' '.join(args)} failed").strip()
        raise PublishError(message)
    return proc


def _require_git_identity():
    name = _git_run(['config', '--get', 'user.name'], allow_failure=True).stdout.strip()
    email = _git_run(['config', '--get', 'user.email'], allow_failure=True).stdout.strip()
    if not name or not email:
        raise PublishError(
            'Git is not fully configured. Set user.name and user.email before publishing.'
        )


def _require_git_remote(branch: str):
    remote = _git_run(['remote', 'get-url', 'origin'], allow_failure=True)
    remote_url = remote.stdout.strip()
    if not remote_url:
        raise PublishError('Git remote "origin" is not configured. Add a GitHub remote before publishing.')

    try:
        _git_run(['rev-parse', '--is-inside-work-tree'])
    except PublishError as exc:
        raise PublishError('This folder is not a Git repository.') from exc

    _git_run(['rev-parse', '--abbrev-ref', 'HEAD'], allow_failure=True)
    logger.info('Publishing to origin/%s (%s)', branch, remote_url)


def _row_to_product(row) -> Dict:
    product = dict(row)
    product['image_path'] = get_catalog_image_path(product.get('image'))
    return product


def _row_to_retailer(row) -> Dict:
    return dict(row)


def _row_to_order(row) -> Dict:
    order = dict(row)
    retailer = get_retailer(order.get('retailer_id', 0)) if order.get('retailer_id') else None
    order['retailer_name'] = retailer.get('shop_name', '') if retailer else ''
    return order


def list_products(
    search: str = '',
    category: str = '',
    supplier: str = '',
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    min_stock: Optional[float] = None,
    max_stock: Optional[float] = None,
    gst: Optional[float] = None,
    barcode: str = '',
    stock_mode: str = 'all',
    status: str = 'all',
) -> List[Dict]:
    connection = get_connection()
    try:
        query = 'SELECT * FROM products WHERE 1=1'
        params: List = []
        if search:
            query += ' AND (product_name LIKE ? OR category LIKE ? OR barcode LIKE ? OR CAST(selling_price AS TEXT) LIKE ? OR CAST(price AS TEXT) LIKE ?)'
            pattern = f'%{search}%'
            params.extend([pattern, pattern, pattern, pattern, pattern])
        if category and category.lower() != 'all':
            query += ' AND category = ?'
            params.append(category)
        if supplier and supplier.lower() != 'all':
            query += ' AND supplier = ?'
            params.append(supplier)
        if min_price is not None:
            query += ' AND selling_price >= ?'
            params.append(float(min_price))
        if max_price is not None and max_price > 0:
            query += ' AND selling_price <= ?'
            params.append(float(max_price))
        if min_stock is not None:
            query += ' AND stock >= ?'
            params.append(float(min_stock))
        if max_stock is not None and max_stock > 0:
            query += ' AND stock <= ?'
            params.append(float(max_stock))
        if gst is not None:
            query += ' AND gst = ?'
            params.append(float(gst))
        if barcode:
            query += ' AND barcode LIKE ?'
            params.append(f'%{barcode.strip()}%')
        if stock_mode == 'out':
            query += ' AND stock <= 0'
        elif stock_mode == 'low':
            query += ' AND stock > 0 AND stock <= min_stock'
        elif stock_mode == 'in':
            query += ' AND stock > 0'
        if status and status.lower() != 'all':
            query += ' AND status = ?'
            params.append('Inactive' if status.lower() == 'inactive' else 'Active')
        query += ' ORDER BY product_name COLLATE NOCASE ASC'
        rows = connection.execute(query, params).fetchall()
        return [_row_to_product(row) for row in rows]
    finally:
        connection.close()


def _normalize_retailer_payload(payload: Dict) -> Dict:
    return {
        'shop_name': (payload.get('shop_name') or '').strip(),
        'owner_name': (payload.get('owner_name') or '').strip(),
        'mobile_number': (payload.get('mobile_number') or '').strip(),
        'whatsapp_number': (payload.get('whatsapp_number') or '').strip(),
        'address': (payload.get('address') or '').strip(),
        'city': (payload.get('city') or '').strip(),
        'gst_number': (payload.get('gst_number') or '').strip(),
        'credit_limit': _safe_float(payload.get('credit_limit', 0)),
        'outstanding_balance': _safe_float(payload.get('outstanding_balance', 0)),
        'status': 'Inactive' if str(payload.get('status') or 'Active').strip().lower() == 'inactive' else 'Active',
    }


def list_retailers(search: str = '', status: str = 'all', city: str = '') -> List[Dict]:
    connection = get_connection()
    try:
        query = 'SELECT * FROM retailers WHERE 1=1'
        params: List = []
        if search:
            query += (
                ' AND (shop_name LIKE ? OR owner_name LIKE ? OR mobile_number LIKE ? OR '
                'whatsapp_number LIKE ? OR address LIKE ? OR city LIKE ? OR gst_number LIKE ?)'
            )
            pattern = f'%{search}%'
            params.extend([pattern] * 7)
        if status and status.lower() != 'all':
            query += ' AND status = ?'
            params.append('Inactive' if status.lower() == 'inactive' else 'Active')
        if city and city.lower() != 'all':
            query += ' AND city = ?'
            params.append(city)
        query += ' ORDER BY shop_name COLLATE NOCASE ASC'
        rows = connection.execute(query, params).fetchall()
        return [_row_to_retailer(row) for row in rows]
    finally:
        connection.close()


def list_products_lookup() -> List[Dict]:
    connection = get_connection()
    try:
        rows = connection.execute(
            "SELECT id, product_name, wholesale_price, selling_price, price FROM products WHERE COALESCE(status, 'Active') = 'Active' ORDER BY product_name COLLATE NOCASE ASC"
        ).fetchall()
        products = []
        for row in rows:
            product = dict(row)
            product['default_wholesale_price'] = _safe_float(product.get('wholesale_price', 0))
            product['default_selling_price'] = _safe_float(product.get('selling_price', product.get('price', 0)))
            products.append(product)
        return products
    finally:
        connection.close()


def _generate_order_number(connection) -> str:
    year = datetime.now().strftime('%Y')
    prefix = f'SO-{year}-'
    row = connection.execute(
        'SELECT order_number FROM orders WHERE order_number LIKE ? ORDER BY id DESC LIMIT 1',
        (f'{prefix}%',),
    ).fetchone()
    next_index = 1
    if row and row['order_number']:
        suffix = row['order_number'].split('-')[-1]
        if suffix.isdigit():
            next_index = int(suffix) + 1
    return f'{prefix}{next_index:05d}'


def _normalize_order_status(status: str) -> str:
    normalized = (status or 'Pending').strip().lower()
    mapping = {
        'pending': 'Pending',
        'confirmed': 'Confirmed',
        'dispatched': 'Dispatched',
        'delivered': 'Delivered',
    }
    return mapping.get(normalized, 'Pending')


def list_orders(search: str = '', status: str = 'all', retailer: str = '') -> List[Dict]:
    connection = get_connection()
    try:
        query = '''
            SELECT o.*, r.shop_name AS retailer_name, r.owner_name AS retailer_owner
            FROM orders o
            LEFT JOIN retailers r ON r.id = o.retailer_id
            WHERE 1=1
        '''
        params: List = []
        if search:
            query += ' AND (o.order_number LIKE ? OR r.shop_name LIKE ? OR r.owner_name LIKE ? OR o.notes LIKE ?)'
            pattern = f'%{search}%'
            params.extend([pattern, pattern, pattern, pattern]
            )
        if status and status.lower() != 'all':
            query += ' AND o.status = ?'
            params.append(_normalize_order_status(status))
        if retailer and retailer.lower() != 'all':
            query += ' AND r.shop_name = ?'
            params.append(retailer)
        query += ' ORDER BY o.order_date DESC, o.id DESC'
        rows = connection.execute(query, params).fetchall()
        orders = []
        for row in rows:
            order = dict(row)
            order['retailer_name'] = order.get('retailer_name', '') or ''
            order['retailer_owner'] = order.get('retailer_owner', '') or ''
            orders.append(order)
        return orders
    finally:
        connection.close()


def get_order(order_id: int) -> Optional[Dict]:
    connection = get_connection()
    try:
        order_row = connection.execute(
            '''
            SELECT o.*, r.shop_name AS retailer_name, r.owner_name AS retailer_owner
            FROM orders o
            LEFT JOIN retailers r ON r.id = o.retailer_id
            WHERE o.id = ?
            ''',
            (order_id,),
        ).fetchone()
        if not order_row:
            return None
        order = dict(order_row)
        order['retailer_name'] = order.get('retailer_name', '') or ''
        order['retailer_owner'] = order.get('retailer_owner', '') or ''
        items = connection.execute(
            '''
            SELECT oi.*, p.product_name AS live_product_name
            FROM order_items oi
            LEFT JOIN products p ON p.id = oi.product_id
            WHERE oi.order_id = ?
            ORDER BY oi.id ASC
            ''',
            (order_id,),
        ).fetchall()
        order['items'] = [dict(item) for item in items]
        return order
    finally:
        connection.close()


def save_order(payload: Dict, items: List[Dict], order_id: Optional[int] = None) -> int:
    if not items:
        raise ValueError('At least one product is required.')

    retailer_id = int(payload.get('retailer_id') or 0)
    if retailer_id <= 0:
        raise ValueError('Retailer is required.')

    order_date = (payload.get('order_date') or datetime.now().date().isoformat()).strip()
    status = _normalize_order_status(payload.get('status', 'Pending'))
    notes = (payload.get('notes') or '').strip()

    normalized_items = []
    total_amount = 0.0
    for item in items:
        product_id = int(item.get('product_id') or 0)
        quantity = _safe_float(item.get('quantity'))
        wholesale_price = _safe_float(item.get('wholesale_price'))
        if product_id <= 0 or quantity <= 0:
            continue
        product_name = (item.get('product_name') or '').strip()
        line_total = round(quantity * wholesale_price, 2)
        total_amount += line_total
        normalized_items.append({
            'product_id': product_id,
            'product_name': product_name,
            'quantity': quantity,
            'wholesale_price': wholesale_price,
            'line_total': line_total,
        })

    if not normalized_items:
        raise ValueError('At least one valid product row is required.')

    now = datetime.now().isoformat(timespec='seconds')
    connection = get_connection()
    try:
        if order_id is None:
            order_number = _generate_order_number(connection)
            cursor = connection.execute(
                '''
                INSERT INTO orders (
                    retailer_id, order_number, order_date, status, total_amount, notes,
                    created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (retailer_id, order_number, order_date, status, total_amount, notes, now, now),
            )
            order_id = int(cursor.lastrowid)
        else:
            order_number = connection.execute('SELECT order_number FROM orders WHERE id = ?', (order_id,)).fetchone()
            if not order_number:
                raise ValueError('Order not found.')
            connection.execute(
                '''
                UPDATE orders
                SET retailer_id = ?, order_date = ?, status = ?, total_amount = ?, notes = ?, updated_at = ?
                WHERE id = ?
                ''',
                (retailer_id, order_date, status, total_amount, notes, now, order_id),
            )
            connection.execute('DELETE FROM order_items WHERE order_id = ?', (order_id,))

        for item in normalized_items:
            connection.execute(
                '''
                INSERT INTO order_items (
                    order_id, product_id, product_name, quantity, wholesale_price, line_total
                ) VALUES (?, ?, ?, ?, ?, ?)
                ''',
                (
                    order_id,
                    item['product_id'],
                    item['product_name'],
                    item['quantity'],
                    item['wholesale_price'],
                    item['line_total'],
                ),
            )
        connection.commit()
        return int(order_id)
    finally:
        connection.close()


def delete_order(order_id: int):
    connection = get_connection()
    try:
        connection.execute('DELETE FROM order_items WHERE order_id = ?', (order_id,))
        connection.execute('DELETE FROM orders WHERE id = ?', (order_id,))
        connection.commit()
    finally:
        connection.close()


def _order_export_rows() -> List[Dict]:
    orders = list_orders()
    rows = []
    for order in orders:
        rows.append({
            'Order Number': order.get('order_number', ''),
            'Order Date': order.get('order_date', ''),
            'Retailer': order.get('retailer_name', ''),
            'Owner Name': order.get('retailer_owner', ''),
            'Status': order.get('status', ''),
            'Total Amount': order.get('total_amount', 0),
            'Notes': order.get('notes', ''),
        })
    return rows


def export_orders_excel(file_path: str):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Orders'
    rows = _order_export_rows()
    headers = list(rows[0].keys()) if rows else ['Order Number', 'Order Date', 'Retailer', 'Owner Name', 'Status', 'Total Amount', 'Notes']
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, '') for header in headers])
    workbook.save(file_path)


def _order_pdf_lines(order_id: int) -> List[str]:
    order = get_order(order_id)
    if not order:
        raise ValueError('Order not found.')
    lines = [
        f"Order Number: {order.get('order_number', '')}",
        f"Order Date: {order.get('order_date', '')}",
        f"Retailer: {order.get('retailer_name', '')}",
        f"Owner: {order.get('retailer_owner', '')}",
        f"Status: {order.get('status', '')}",
        '',
        'Items:',
    ]
    for item in order.get('items', []):
        lines.append(
            f"- {item.get('product_name', '')} x {float(item.get('quantity', 0) or 0):.2f} @ {float(item.get('wholesale_price', 0) or 0):.2f} = {float(item.get('line_total', 0) or 0):.2f}"
        )
    lines.extend(['', f"Total Amount: {float(order.get('total_amount', 0) or 0):.2f}"])
    if order.get('notes'):
        lines.extend(['', f"Notes: {order.get('notes', '')}"])
    return lines


def export_order_pdf(order_id: int, file_path: str):
    from PySide6 import QtCore, QtGui

    document = QtGui.QTextDocument()
    order = get_order(order_id)
    if not order:
        raise ValueError('Order not found.')
    item_rows = []
    for item in order.get('items', []):
        item_rows.append(
            '<tr>'
            f'<td>{item.get("product_name", "")}</td>'
            f'<td style="text-align:right;">{float(item.get("quantity", 0) or 0):.2f}</td>'
            f'<td style="text-align:right;">{float(item.get("wholesale_price", 0) or 0):.2f}</td>'
            f'<td style="text-align:right;">{float(item.get("line_total", 0) or 0):.2f}</td>'
            '</tr>'
        )
    html = f'''
    <html>
      <body style="font-family: Arial, sans-serif;">
        <h2>Sales Order</h2>
        <p><strong>Order Number:</strong> {order.get('order_number', '')}<br>
        <strong>Order Date:</strong> {order.get('order_date', '')}<br>
        <strong>Retailer:</strong> {order.get('retailer_name', '')}<br>
        <strong>Owner:</strong> {order.get('retailer_owner', '')}<br>
        <strong>Status:</strong> {order.get('status', '')}</p>
        <table cellspacing="0" cellpadding="6" border="1" width="100%">
          <thead>
            <tr><th align="left">Product</th><th align="right">Qty</th><th align="right">Wholesale</th><th align="right">Total</th></tr>
          </thead>
          <tbody>
            {''.join(item_rows)}
          </tbody>
        </table>
        <p><strong>Total Amount:</strong> {float(order.get('total_amount', 0) or 0):.2f}</p>
        <p><strong>Notes:</strong> {order.get('notes', '')}</p>
      </body>
    </html>
    '''
    document.setHtml(html)
    writer = QtGui.QPdfWriter(str(file_path))
    writer.setPageMargins(QtCore.QMarginsF(10, 10, 10, 10))
    document.print_(writer)


def print_order(order_id: int, printer):
    from PySide6 import QtGui

    document = QtGui.QTextDocument()
    lines = _order_pdf_lines(order_id)
    html = '<br>'.join(line.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;') for line in lines)
    document.setHtml(f'<html><body style="font-family: Arial, sans-serif; white-space: pre-wrap;">{html}</body></html>')
    document.print_(printer)


def get_retailer(retailer_id: int) -> Optional[Dict]:
    connection = get_connection()
    try:
        row = connection.execute('SELECT * FROM retailers WHERE id = ?', (retailer_id,)).fetchone()
        return _row_to_retailer(row) if row else None
    finally:
        connection.close()


def create_retailer(payload: Dict):
    data = _normalize_retailer_payload(payload)
    if not data['shop_name']:
        raise ValueError('Shop Name is required.')

    now = datetime.now().isoformat(timespec='seconds')
    connection = get_connection()
    try:
        connection.execute(
            '''
            INSERT INTO retailers (
                shop_name, owner_name, mobile_number, whatsapp_number, address,
                city, gst_number, credit_limit, outstanding_balance, status,
                created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                data['shop_name'],
                data['owner_name'],
                data['mobile_number'],
                data['whatsapp_number'],
                data['address'],
                data['city'],
                data['gst_number'],
                data['credit_limit'],
                data['outstanding_balance'],
                data['status'],
                now,
                now,
            ),
        )
        connection.commit()
    finally:
        connection.close()


def update_retailer(retailer_id: int, payload: Dict):
    data = _normalize_retailer_payload(payload)
    if not data['shop_name']:
        raise ValueError('Shop Name is required.')

    now = datetime.now().isoformat(timespec='seconds')
    connection = get_connection()
    try:
        existing = connection.execute('SELECT id FROM retailers WHERE id = ?', (retailer_id,)).fetchone()
        if not existing:
            raise ValueError('Retailer not found.')
        connection.execute(
            '''
            UPDATE retailers
            SET shop_name = ?, owner_name = ?, mobile_number = ?, whatsapp_number = ?, address = ?,
                city = ?, gst_number = ?, credit_limit = ?, outstanding_balance = ?, status = ?,
                updated_at = ?
            WHERE id = ?
            ''',
            (
                data['shop_name'],
                data['owner_name'],
                data['mobile_number'],
                data['whatsapp_number'],
                data['address'],
                data['city'],
                data['gst_number'],
                data['credit_limit'],
                data['outstanding_balance'],
                data['status'],
                now,
                retailer_id,
            ),
        )
        connection.commit()
    finally:
        connection.close()


def delete_retailer(retailer_id: int):
    connection = get_connection()
    try:
        connection.execute('DELETE FROM retailers WHERE id = ?', (retailer_id,))
        connection.commit()
    finally:
        connection.close()


def _normalize_retailer_header(header: str) -> str:
    import re

    return re.sub(r'[^a-z0-9]+', '', (header or '').strip().lower())


def import_retailers_from_excel(file_path: str) -> ImportReport:
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers_raw = [str(cell).strip() if cell is not None else '' for cell in header_row]
    lookup = {_normalize_retailer_header(name): idx for idx, name in enumerate(headers_raw)}

    aliases = {
        'shop_name': ['shopname', 'shop', 'storename'],
        'owner_name': ['ownername', 'owner'],
        'mobile_number': ['mobile', 'mobilenumber', 'phone'],
        'whatsapp_number': ['whatsapp', 'whatsappnumber'],
        'address': ['address', 'location'],
        'city': ['city', 'town'],
        'gst_number': ['gst', 'gstnumber', 'gstno'],
        'credit_limit': ['creditlimit', 'limit'],
        'outstanding_balance': ['outstandingbalance', 'balance', 'due'],
        'status': ['status', 'activeinactive', 'state'],
    }

    def index_of(field: str) -> Optional[int]:
        for alias in aliases[field]:
            if alias in lookup:
                return lookup[alias]
        return None

    idx = {field: index_of(field) for field in aliases}
    if idx['shop_name'] is None:
        raise ValueError('Import file must contain Shop Name column.')

    connection = get_connection()
    report = ImportReport()
    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            try:
                shop_name = str(row[idx['shop_name']]).strip() if idx['shop_name'] is not None and row[idx['shop_name']] is not None else ''
                if not shop_name:
                    report.skipped += 1
                    continue

                payload = {
                    'shop_name': shop_name,
                    'owner_name': str(row[idx['owner_name']]).strip() if idx['owner_name'] is not None and row[idx['owner_name']] is not None else '',
                    'mobile_number': str(row[idx['mobile_number']]).strip() if idx['mobile_number'] is not None and row[idx['mobile_number']] is not None else '',
                    'whatsapp_number': str(row[idx['whatsapp_number']]).strip() if idx['whatsapp_number'] is not None and row[idx['whatsapp_number']] is not None else '',
                    'address': str(row[idx['address']]).strip() if idx['address'] is not None and row[idx['address']] is not None else '',
                    'city': str(row[idx['city']]).strip() if idx['city'] is not None and row[idx['city']] is not None else '',
                    'gst_number': str(row[idx['gst_number']]).strip() if idx['gst_number'] is not None and row[idx['gst_number']] is not None else '',
                    'credit_limit': _safe_float(row[idx['credit_limit']]) if idx['credit_limit'] is not None else 0,
                    'outstanding_balance': _safe_float(row[idx['outstanding_balance']]) if idx['outstanding_balance'] is not None else 0,
                    'status': str(row[idx['status']]).strip() if idx['status'] is not None and row[idx['status']] is not None else 'Active',
                }

                existing = None
                if payload['mobile_number']:
                    existing = connection.execute(
                        'SELECT id FROM retailers WHERE mobile_number = ?',
                        (payload['mobile_number'],),
                    ).fetchone()
                if existing is None:
                    existing = connection.execute(
                        'SELECT id FROM retailers WHERE LOWER(shop_name) = LOWER(?)',
                        (payload['shop_name'],),
                    ).fetchone()

                data = _normalize_retailer_payload(payload)
                if existing:
                    connection.execute(
                        '''
                        UPDATE retailers
                        SET shop_name = ?, owner_name = ?, mobile_number = ?, whatsapp_number = ?, address = ?,
                            city = ?, gst_number = ?, credit_limit = ?, outstanding_balance = ?, status = ?,
                            updated_at = ?
                        WHERE id = ?
                        ''',
                        (
                            data['shop_name'],
                            data['owner_name'],
                            data['mobile_number'],
                            data['whatsapp_number'],
                            data['address'],
                            data['city'],
                            data['gst_number'],
                            data['credit_limit'],
                            data['outstanding_balance'],
                            data['status'],
                            datetime.now().isoformat(timespec='seconds'),
                            existing['id'],
                        ),
                    )
                    report.updated += 1
                else:
                    now = datetime.now().isoformat(timespec='seconds')
                    connection.execute(
                        '''
                        INSERT INTO retailers (
                            shop_name, owner_name, mobile_number, whatsapp_number, address,
                            city, gst_number, credit_limit, outstanding_balance, status,
                            created_at, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''',
                        (
                            data['shop_name'],
                            data['owner_name'],
                            data['mobile_number'],
                            data['whatsapp_number'],
                            data['address'],
                            data['city'],
                            data['gst_number'],
                            data['credit_limit'],
                            data['outstanding_balance'],
                            data['status'],
                            now,
                            now,
                        ),
                    )
                    report.added += 1
            except Exception:
                report.errors += 1
                logger.exception('Retailer import row failed: %s', row)

        connection.commit()
    finally:
        connection.close()

    return report


def _retailer_export_dict(retailer: Dict) -> Dict:
    return {
        'Shop Name': retailer.get('shop_name', ''),
        'Owner Name': retailer.get('owner_name', ''),
        'Mobile Number': retailer.get('mobile_number', ''),
        'WhatsApp Number': retailer.get('whatsapp_number', ''),
        'Address': retailer.get('address', ''),
        'City': retailer.get('city', ''),
        'GST Number': retailer.get('gst_number', ''),
        'Credit Limit': retailer.get('credit_limit', 0),
        'Outstanding Balance': retailer.get('outstanding_balance', 0),
        'Status': retailer.get('status', 'Active'),
    }


def export_retailers_excel(file_path: str):
    retailers = list_retailers()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Retailers'
    headers = list(_retailer_export_dict({}).keys())
    sheet.append(headers)
    for retailer in retailers:
        row_dict = _retailer_export_dict(retailer)
        sheet.append([row_dict[header] for header in headers])
    workbook.save(file_path)


def get_product(product_id: int) -> Optional[Dict]:
    connection = get_connection()
    try:
        row = connection.execute('SELECT * FROM products WHERE id = ?', (product_id,)).fetchone()
        return _row_to_product(row) if row else None
    finally:
        connection.close()


def list_categories(search: str = '') -> List[Dict]:
    connection = get_connection()
    try:
        if search:
            rows = connection.execute(
                'SELECT * FROM categories WHERE category_name LIKE ? ORDER BY category_name COLLATE NOCASE ASC',
                (f'%{search}%',),
            ).fetchall()
        else:
            rows = connection.execute(
                'SELECT * FROM categories ORDER BY category_name COLLATE NOCASE ASC'
            ).fetchall()
        return [dict(row) for row in rows]
    finally:
        connection.close()


def ensure_category(name: str):
    category = (name or '').strip()
    if not category:
        return
    connection = get_connection()
    try:
        exists = connection.execute(
            'SELECT id FROM categories WHERE LOWER(category_name) = LOWER(?)', (category,)
        ).fetchone()
        if not exists:
            connection.execute('INSERT INTO categories (category_name) VALUES (?)', (category,))
            connection.commit()
    finally:
        connection.close()


def create_category(name: str):
    ensure_category(name)
    return list_categories()


def rename_category(category_id: int, new_name: str):
    connection = get_connection()
    try:
        row = connection.execute('SELECT category_name FROM categories WHERE id = ?', (category_id,)).fetchone()
        if not row:
            return list_categories()
        old_name = row['category_name']
        new_name_clean = (new_name or '').strip()
        if not new_name_clean:
            return list_categories()
        connection.execute('UPDATE categories SET category_name = ? WHERE id = ?', (new_name_clean, category_id))
        connection.execute('UPDATE products SET category = ? WHERE category = ?', (new_name_clean, old_name))
        connection.commit()
        return list_categories()
    finally:
        connection.close()


def delete_category(category_id: int):
    connection = get_connection()
    try:
        row = connection.execute('SELECT category_name FROM categories WHERE id = ?', (category_id,)).fetchone()
        if not row:
            return list_categories()
        in_use = connection.execute(
            'SELECT COUNT(*) AS cnt FROM products WHERE category = ?', (row['category_name'],)
        ).fetchone()['cnt']
        if in_use > 0:
            raise ValueError('Cannot delete category in use by products.')
        connection.execute('DELETE FROM categories WHERE id = ?', (category_id,))
        connection.commit()
        return list_categories()
    finally:
        connection.close()


def _normalize_payload(payload: Dict) -> Dict:
    selling_price = float(payload.get('selling_price', payload.get('price', 0)) or 0)
    purchase_price = float(payload.get('purchase_price', 0) or 0)
    wholesale_price = float(payload.get('wholesale_price', 0) or 0)
    return {
        'product_name': (payload.get('product_name') or '').strip(),
        'category': (payload.get('category') or 'General').strip() or 'General',
        'purchase_price': purchase_price,
        'selling_price': selling_price,
        'wholesale_price': wholesale_price,
        'barcode': (payload.get('barcode') or '').strip(),
        'supplier': (payload.get('supplier') or '').strip(),
        'description': (payload.get('description') or '').strip(),
        'gst': float(payload.get('gst', 0) or 0),
        'hsn': (payload.get('hsn') or '').strip(),
        'stock': float(payload.get('stock', 0) or 0),
        'min_stock': float(payload.get('min_stock', 0) or 0),
        'status': 'Inactive' if str(payload.get('status') or 'Active').strip().lower() == 'inactive' else 'Active',
        'image': normalize_image_value(payload.get('image')),
        'price': selling_price,
    }


def _ensure_unique_barcode(connection, barcode: str, exclude_id: Optional[int] = None):
    barcode_clean = (barcode or '').strip()
    if not barcode_clean:
        return

    query = 'SELECT id, product_name FROM products WHERE barcode = ?'
    params: List = [barcode_clean]
    if exclude_id is not None:
        query += ' AND id != ?'
        params.append(exclude_id)
    row = connection.execute(query, params).fetchone()
    if row:
        raise ValueError(f"Barcode already exists for product: {row['product_name']}")


def create_product(payload: Dict, image_source: str = ''):
    data = _normalize_payload(payload)
    ensure_category(data['category'])
    if image_source:
        data['image'] = save_image_from_path(image_source, data['product_name'])

    now = datetime.now().isoformat(timespec='seconds')
    connection = get_connection()
    try:
        _ensure_unique_barcode(connection, data['barcode'])
        connection.execute(
            '''
            INSERT INTO products (
                product_name, category, price, description, image,
                purchase_price, selling_price, wholesale_price,
                barcode, gst, hsn, stock, min_stock, status, supplier,
                created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                data['product_name'],
                data['category'],
                data['price'],
                data['description'],
                data['image'],
                data['purchase_price'],
                data['selling_price'],
                data['wholesale_price'],
                data['barcode'],
                data['gst'],
                data['hsn'],
                data['stock'],
                data['min_stock'],
                data['status'],
                data['supplier'],
                now,
                now,
            ),
        )
        connection.commit()
    finally:
        connection.close()

    generate_catalog_json()


def _delete_image_if_unused(connection, image_name: str, exclude_id: Optional[int] = None):
    image_name = normalize_image_value(image_name)
    if not image_name:
        return
    query = 'SELECT COUNT(*) AS cnt FROM products WHERE image = ?'
    params: List = [image_name]
    if exclude_id is not None:
        query += ' AND id != ?'
        params.append(exclude_id)
    cnt = connection.execute(query, params).fetchone()['cnt']
    if cnt == 0:
        disk_path = IMAGES_DIR / image_name
        if disk_path.exists() and disk_path.is_file():
            disk_path.unlink()


def update_product(product_id: int, payload: Dict, image_source: str = ''):
    data = _normalize_payload(payload)
    ensure_category(data['category'])
    connection = get_connection()
    try:
        existing = connection.execute('SELECT * FROM products WHERE id = ?', (product_id,)).fetchone()
        if not existing:
            raise ValueError('Product not found.')

        _ensure_unique_barcode(connection, data['barcode'], exclude_id=product_id)

        old_image = normalize_image_value(existing['image'])
        data['image'] = data['image'] or old_image
        if image_source:
            data['image'] = save_image_from_path(image_source, data['product_name'])
            _delete_image_if_unused(connection, old_image, exclude_id=product_id)

        connection.execute(
            '''
            UPDATE products
            SET product_name = ?, category = ?, price = ?, description = ?, image = ?,
                purchase_price = ?, selling_price = ?, wholesale_price = ?,
                barcode = ?, gst = ?, hsn = ?, stock = ?, min_stock = ?, status = ?, supplier = ?, updated_at = ?
            WHERE id = ?
            ''',
            (
                data['product_name'],
                data['category'],
                data['price'],
                data['description'],
                data['image'],
                data['purchase_price'],
                data['selling_price'],
                data['wholesale_price'],
                data['barcode'],
                data['gst'],
                data['hsn'],
                data['stock'],
                data['min_stock'],
                data['status'],
                data['supplier'],
                datetime.now().isoformat(timespec='seconds'),
                product_id,
            ),
        )
        connection.commit()
    finally:
        connection.close()

    generate_catalog_json()


def duplicate_product(product_id: int):
    product = get_product(product_id)
    if not product:
        raise ValueError('Product not found.')

    duplicate_name = f"{product['product_name']} Copy"
    payload = {
        'product_name': duplicate_name,
        'category': product['category'],
        'purchase_price': product.get('purchase_price', 0),
        'selling_price': product.get('selling_price', product.get('price', 0)),
        'wholesale_price': product.get('wholesale_price', 0),
        'barcode': '',
        'description': product.get('description', ''),
        'gst': product.get('gst', 0),
        'hsn': product.get('hsn', ''),
        'stock': product.get('stock', 0),
        'min_stock': product.get('min_stock', 0),
        'status': product.get('status', 'Active'),
        'supplier': product.get('supplier', ''),
        'image': product.get('image', ''),
    }
    create_product(payload)


def delete_product(product_id: int):
    connection = get_connection()
    try:
        row = connection.execute('SELECT image FROM products WHERE id = ?', (product_id,)).fetchone()
        if not row:
            return
        image_name = normalize_image_value(row['image'])
        connection.execute('DELETE FROM products WHERE id = ?', (product_id,))
        _delete_image_if_unused(connection, image_name)
        connection.commit()
    finally:
        connection.close()

    generate_catalog_json()


def generate_catalog_json() -> List[Dict]:
    connection = get_connection()
    try:
        rows = connection.execute(
            "SELECT * FROM products WHERE COALESCE(status, 'Active') = 'Active' ORDER BY product_name COLLATE NOCASE ASC"
        ).fetchall()
    finally:
        connection.close()

    products = []
    for row in rows:
        selling_price = row['selling_price'] if row['selling_price'] is not None else row['price']
        products.append(
            {
                'id': slugify(row['product_name']),
                'name': row['product_name'],
                'category': row['category'] or 'General',
                'price': round(float(selling_price or 0), 2),
                'image': get_catalog_image_path(row['image']),
                'description': row['description'] or '',
            }
        )

    target = ROOT_DIR / 'products.json'
    with open(target, 'w', encoding='utf-8') as handle:
        json.dump({'products': products}, handle, indent=2)
        handle.write('\n')

    return products


def _normalize_header(header: str) -> str:
    import re

    return re.sub(r'[^a-z0-9]+', '', (header or '').strip().lower())


def import_products_from_excel(file_path: str) -> ImportReport:
    return import_products_from_excel_batched(file_path)


def _save_import_report(log_rows: List[Dict]) -> str:
    if not log_rows:
        return ''
    report_name = f"Import_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    report_path = ROOT_DIR / report_name
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Import Report'
    sheet.append(['Row Number', 'Product Name', 'Reason'])
    for row in log_rows:
        sheet.append([row.get('row_number', ''), row.get('product_name', ''), row.get('reason', '')])
    workbook.save(report_path)
    return str(report_path)


def import_products_from_excel_batched(
    file_path: str,
    options: Optional[Dict] = None,
    progress_callback: Optional[Callable[[Dict], None]] = None,
    pause_check: Optional[Callable[[], bool]] = None,
    cancel_check: Optional[Callable[[], bool]] = None,
    batch_size: int = 100,
) -> ImportReport:
    options = options or {}
    update_existing = bool(options.get('update_existing', True))
    insert_new = bool(options.get('insert_new', True))
    ignore_duplicate_barcode = bool(options.get('ignore_duplicate_barcode', True))
    ignore_empty_rows = bool(options.get('ignore_empty_rows', True))
    skip_invalid_images = bool(options.get('skip_invalid_images', True))

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    sheet = workbook.active

    headers_raw = [str(cell).strip() if cell is not None else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    lookup = {_normalize_header(name): idx for idx, name in enumerate(headers_raw)}

    aliases = {
        'product_name': ['productname', 'name', 'itemname'],
        'category': ['category', 'group'],
        'supplier': ['supplier', 'brand', 'vendor'],
        'purchase_price': ['purchaseprice', 'buyprice', 'costprice'],
        'selling_price': ['sellingprice', 'price', 'mrp', 'saleprice'],
        'wholesale_price': ['wholesaleprice', 'wholesale'],
        'barcode': ['barcode', 'barcodeno', 'sku'],
        'description': ['description', 'details'],
        'gst': ['gst', 'tax'],
        'hsn': ['hsn', 'hsncode'],
        'stock': ['stock', 'qty', 'quantity'],
        'min_stock': ['minstock', 'minimumstock', 'reorderlevel'],
        'status': ['status', 'activeinactive', 'state'],
        'image': ['image', 'photo', 'photofile', 'imagepath'],
    }

    def index_of(field: str) -> Optional[int]:
        for alias in aliases[field]:
            if alias in lookup:
                return lookup[alias]
        return None

    idx = {field: index_of(field) for field in aliases}
    if idx['product_name'] is None:
        raise ValueError('Import file must contain Product Name column.')

    rows_values = list(sheet.iter_rows(min_row=2, values_only=True))
    total_rows = len(rows_values)
    report = ImportReport(total=total_rows)
    log_rows: List[Dict] = []
    start = time.perf_counter()

    def emit(status: str, imported_count: int, current_product: str = '', percent: int = 0):
        if progress_callback:
            progress_callback(
                {
                    'status': status,
                    'imported': imported_count,
                    'total': total_rows,
                    'current_product': current_product,
                    'percent': percent,
                }
            )

    connection = get_connection()
    connection.execute('PRAGMA temp_store = MEMORY')
    connection.execute('PRAGMA cache_size = -20000')
    connection.execute('PRAGMA synchronous = NORMAL')
    insert_sql = (
        'INSERT INTO products ('
        'product_name, category, price, description, image, purchase_price, selling_price, wholesale_price, '
        'barcode, gst, hsn, stock, min_stock, status, supplier, created_at, updated_at'
        ') VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'
    )
    update_sql = (
        'UPDATE products SET product_name = ?, category = ?, price = ?, description = ?, image = ?, purchase_price = ?, '
        'selling_price = ?, wholesale_price = ?, barcode = ?, gst = ?, hsn = ?, stock = ?, min_stock = ?, status = ?, '
        'supplier = ?, updated_at = ? WHERE id = ?'
    )

    try:
        emit('Reading Excel...', 0, percent=2)
        connection.execute('BEGIN')
        existing_rows = connection.execute('SELECT id, product_name, category, barcode FROM products').fetchall()
        barcode_map: Dict[str, int] = {}
        name_map: Dict[str, int] = {}
        for row in existing_rows:
            barcode_key = (row['barcode'] or '').strip().lower()
            if barcode_key:
                barcode_map[barcode_key] = int(row['id'])
            name_key = f"{(row['product_name'] or '').strip().lower()}|{(row['category'] or '').strip().lower()}"
            if name_key and name_key != '|':
                name_map[name_key] = int(row['id'])

        insert_batch: List[tuple] = []
        update_batch: List[tuple] = []
        processed = 0
        seen_import_barcodes = set()
        seen_import_names = set()

        for row_number, row in enumerate(rows_values, start=2):
            while pause_check and pause_check():
                time.sleep(0.1)
            if cancel_check and cancel_check():
                raise RuntimeError('IMPORT_CANCELLED')

            if ignore_empty_rows and not any(cell is not None and str(cell).strip() for cell in row):
                report.skipped += 1
                log_rows.append({'row_number': row_number, 'product_name': '', 'reason': 'Empty row'})
                continue

            product_name = str(row[idx['product_name']]).strip() if idx['product_name'] is not None and row[idx['product_name']] is not None else ''
            if not product_name:
                report.skipped += 1
                log_rows.append({'row_number': row_number, 'product_name': '', 'reason': 'Missing product name'})
                continue

            try:
                payload = {
                    'product_name': product_name,
                    'category': str(row[idx['category']]).strip() if idx['category'] is not None and row[idx['category']] is not None else 'General',
                    'supplier': str(row[idx['supplier']]).strip() if idx['supplier'] is not None and row[idx['supplier']] is not None else '',
                    'purchase_price': _safe_float(row[idx['purchase_price']]) if idx['purchase_price'] is not None else 0,
                    'selling_price': _safe_float(row[idx['selling_price']]) if idx['selling_price'] is not None else 0,
                    'wholesale_price': _safe_float(row[idx['wholesale_price']]) if idx['wholesale_price'] is not None else 0,
                    'barcode': str(row[idx['barcode']]).strip() if idx['barcode'] is not None and row[idx['barcode']] is not None else '',
                    'description': str(row[idx['description']]).strip() if idx['description'] is not None and row[idx['description']] is not None else '',
                    'gst': _safe_float(row[idx['gst']]) if idx['gst'] is not None else 0,
                    'hsn': str(row[idx['hsn']]).strip() if idx['hsn'] is not None and row[idx['hsn']] is not None else '',
                    'stock': _safe_float(row[idx['stock']]) if idx['stock'] is not None else 0,
                    'min_stock': _safe_float(row[idx['min_stock']]) if idx['min_stock'] is not None else 0,
                    'status': str(row[idx['status']]).strip() if idx['status'] is not None and row[idx['status']] is not None else 'Active',
                    'image': str(row[idx['image']]).strip() if idx['image'] is not None and row[idx['image']] is not None else '',
                }

                if skip_invalid_images and payload.get('image'):
                    image_name = normalize_image_value(payload['image'])
                    if image_name and not (ROOT_DIR / 'images' / image_name).exists():
                        payload['image'] = ''

                ensure_category(payload['category'])
                data = _normalize_payload(payload)
                barcode_key = data['barcode'].strip().lower()
                name_key = f"{data['product_name'].strip().lower()}|{data['category'].strip().lower()}"

                if barcode_key and barcode_key in seen_import_barcodes:
                    report.duplicate += 1
                    reason = 'Duplicate barcode in import file'
                    if ignore_duplicate_barcode:
                        report.skipped += 1
                        log_rows.append({'row_number': row_number, 'product_name': data['product_name'], 'reason': reason})
                    else:
                        report.errors += 1
                        log_rows.append({'row_number': row_number, 'product_name': data['product_name'], 'reason': reason})
                    continue

                if name_key in seen_import_names:
                    report.duplicate += 1
                    report.skipped += 1
                    log_rows.append({'row_number': row_number, 'product_name': data['product_name'], 'reason': 'Duplicate product name in import file'})
                    continue

                existing_id = barcode_map.get(barcode_key) if barcode_key else None
                if existing_id is None:
                    existing_id = name_map.get(name_key)

                seen_import_barcodes.add(barcode_key)
                seen_import_names.add(name_key)

                if existing_id:
                    if not update_existing:
                        report.skipped += 1
                        log_rows.append({'row_number': row_number, 'product_name': data['product_name'], 'reason': 'Update existing disabled'})
                        continue
                    update_batch.append(
                        (
                            data['product_name'],
                            data['category'],
                            data['price'],
                            data['description'],
                            data['image'],
                            data['purchase_price'],
                            data['selling_price'],
                            data['wholesale_price'],
                            data['barcode'],
                            data['gst'],
                            data['hsn'],
                            data['stock'],
                            data['min_stock'],
                            data['status'],
                            data['supplier'],
                            datetime.now().isoformat(timespec='seconds'),
                            existing_id,
                        )
                    )
                    report.updated += 1
                else:
                    if not insert_new:
                        report.skipped += 1
                        log_rows.append({'row_number': row_number, 'product_name': data['product_name'], 'reason': 'Insert new disabled'})
                        continue
                    now = datetime.now().isoformat(timespec='seconds')
                    insert_batch.append(
                        (
                            data['product_name'],
                            data['category'],
                            data['price'],
                            data['description'],
                            data['image'],
                            data['purchase_price'],
                            data['selling_price'],
                            data['wholesale_price'],
                            data['barcode'],
                            data['gst'],
                            data['hsn'],
                            data['stock'],
                            data['min_stock'],
                            data['status'],
                            data['supplier'],
                            now,
                            now,
                        )
                    )
                    report.added += 1
                    if barcode_key:
                        barcode_map[barcode_key] = -1
                    name_map[name_key] = -1

                processed += 1
                elapsed = max(time.perf_counter() - start, 0.001)
                speed = processed / elapsed
                percent = min(90, int((row_number - 1) * 90 / max(total_rows, 1)))
                emit('Validating...', processed, data['product_name'], percent)
                if progress_callback:
                    progress_callback({'speed': speed})

                if len(insert_batch) + len(update_batch) >= batch_size:
                    emit('Saving Database...', processed, data['product_name'], max(percent, 91))
                    if insert_batch:
                        connection.executemany(insert_sql, insert_batch)
                        insert_batch.clear()
                    if update_batch:
                        connection.executemany(update_sql, update_batch)
                        update_batch.clear()
            except Exception as ex:
                report.errors += 1
                log_rows.append({'row_number': row_number, 'product_name': product_name, 'reason': str(ex)})

        emit('Saving Database...', processed, percent=95)
        if insert_batch:
            connection.executemany(insert_sql, insert_batch)
        if update_batch:
            connection.executemany(update_sql, update_batch)
        connection.commit()

        emit('Generating Catalog...', processed, percent=98)
        generate_catalog_json()
    except Exception as ex:
        connection.rollback()
        if str(ex) != 'IMPORT_CANCELLED':
            raise
    finally:
        connection.close()

    report.elapsed_seconds = time.perf_counter() - start
    report.report_path = _save_import_report(log_rows)
    return report


def _product_export_dict(product: Dict) -> Dict:
    return {
        'Product Name': product.get('product_name', ''),
        'Category': product.get('category', ''),
        'Supplier': product.get('supplier', ''),
        'Purchase Price': product.get('purchase_price', 0),
        'Selling Price': product.get('selling_price', product.get('price', 0)),
        'Wholesale Price': product.get('wholesale_price', 0),
        'Barcode': product.get('barcode', ''),
        'Description': product.get('description', ''),
        'GST': product.get('gst', 0),
        'HSN': product.get('hsn', ''),
        'Stock': product.get('stock', 0),
        'Minimum Stock': product.get('min_stock', 0),
        'Status': product.get('status', 'Active'),
        'Photo': get_catalog_image_path(product.get('image')),
    }


def generate_product_import_sample(file_path: str):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Products Import Sample'

    headers = [
        'Supplier',
        'Product Name',
        'Category',
        'Purchase Price',
        'Selling Price',
        'Wholesale Price',
        'Barcode',
        'GST',
        'HSN',
        'Stock',
        'Minimum Stock',
        'Status',
        'Image Path',
        'Description',
    ]
    sample_row = [
        'Patel Distributors',
        'Sample Product',
        'General',
        120.0,
        150.0,
        130.0,
        '1234567890123',
        18.0,
        '3923',
        25,
        5,
        'Active',
        'images/placeholder.svg',
        'Sample description',
    ]
    sample_row_two = [
        'Metro Supplies',
        'Steel Lunch Box',
        'Kitchen',
        180.0,
        240.0,
        205.0,
        '7894561230001',
        12.0,
        '7323',
        40,
        8,
        'Active',
        'images/lunch-box.jpg',
        'Food grade steel with lock lid',
    ]

    sheet.append(headers)
    sheet.append(sample_row)
    sheet.append(sample_row_two)

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:N1'

    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    widths = [20, 26, 20, 14, 14, 16, 18, 10, 12, 10, 14, 12, 24, 38]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    status_validation = DataValidation(type='list', formula1='"Active,Inactive"', allow_blank=True)
    sheet.add_data_validation(status_validation)
    status_validation.add('L2:L10000')

    workbook.save(file_path)


def update_products_status(product_ids: List[int], status: str):
    valid_status = 'Inactive' if str(status).strip().lower() == 'inactive' else 'Active'
    ids = [int(pid) for pid in product_ids if int(pid) > 0]
    if not ids:
        return

    connection = get_connection()
    try:
        placeholders = ','.join(['?'] * len(ids))
        params: List = [valid_status, datetime.now().isoformat(timespec='seconds'), *ids]
        connection.execute(
            f'UPDATE products SET status = ?, updated_at = ? WHERE id IN ({placeholders})',
            params,
        )
        connection.commit()
    finally:
        connection.close()

    generate_catalog_json()


def list_suppliers(search: str = '') -> List[str]:
    connection = get_connection()
    try:
        query = "SELECT DISTINCT supplier FROM products WHERE TRIM(COALESCE(supplier, '')) <> ''"
        params: List = []
        if search:
            query += ' AND supplier LIKE ?'
            params.append(f'%{search.strip()}%')
        query += ' ORDER BY supplier COLLATE NOCASE ASC'
        rows = connection.execute(query, params).fetchall()
        return [str(row['supplier']).strip() for row in rows if str(row['supplier']).strip()]
    finally:
        connection.close()


def bulk_update_products(product_ids: List[int], updates: Dict):
    ids = [int(pid) for pid in product_ids if int(pid) > 0]
    if not ids:
        return

    allowed_fields = {
        'category',
        'supplier',
        'purchase_price',
        'selling_price',
        'wholesale_price',
        'gst',
        'stock',
        'min_stock',
        'image',
        'status',
    }
    sanitized: Dict = {}
    for key, value in (updates or {}).items():
        if key not in allowed_fields:
            continue
        if key in {'purchase_price', 'selling_price', 'wholesale_price', 'gst', 'stock', 'min_stock'}:
            sanitized[key] = _safe_float(value, 0)
        elif key == 'status':
            sanitized[key] = 'Inactive' if str(value).strip().lower() == 'inactive' else 'Active'
        elif key == 'category':
            sanitized[key] = (value or 'General').strip() or 'General'
        elif key == 'supplier':
            sanitized[key] = (value or '').strip()
        elif key == 'image':
            sanitized[key] = normalize_image_value(value)
        else:
            sanitized[key] = value

    if not sanitized:
        return

    if 'category' in sanitized:
        ensure_category(sanitized['category'])

    connection = get_connection()
    try:
        placeholders = ','.join(['?'] * len(ids))
        set_parts = [f"{field} = ?" for field in sanitized.keys()]
        params: List = list(sanitized.values())
        params.append(datetime.now().isoformat(timespec='seconds'))
        set_parts.append('updated_at = ?')
        params.extend(ids)
        connection.execute(
            f"UPDATE products SET {', '.join(set_parts)} WHERE id IN ({placeholders})",
            params,
        )
        connection.commit()
    finally:
        connection.close()

    generate_catalog_json()


def bulk_delete_products(product_ids: List[int]):
    ids = [int(pid) for pid in product_ids if int(pid) > 0]
    if not ids:
        return
    connection = get_connection()
    try:
        placeholders = ','.join(['?'] * len(ids))
        rows = connection.execute(
            f'SELECT image FROM products WHERE id IN ({placeholders})',
            ids,
        ).fetchall()
        connection.execute(f'DELETE FROM products WHERE id IN ({placeholders})', ids)
        for row in rows:
            _delete_image_if_unused(connection, row['image'])
        connection.commit()
    finally:
        connection.close()

    generate_catalog_json()


def get_product_statistics() -> Dict:
    connection = get_connection()
    try:
        row = connection.execute(
            '''
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN COALESCE(status, 'Active') = 'Active' THEN 1 ELSE 0 END) AS active,
                SUM(CASE WHEN COALESCE(status, 'Active') = 'Inactive' THEN 1 ELSE 0 END) AS inactive,
                SUM(CASE WHEN COALESCE(stock, 0) <= 0 THEN 1 ELSE 0 END) AS out_of_stock,
                SUM(CASE WHEN COALESCE(stock, 0) > 0 AND COALESCE(stock, 0) <= COALESCE(min_stock, 0) THEN 1 ELSE 0 END) AS low_stock,
                SUM(COALESCE(stock, 0) * COALESCE(purchase_price, 0)) AS inventory_value,
                AVG(CASE
                    WHEN COALESCE(selling_price, 0) > 0
                    THEN ((COALESCE(selling_price, 0) - COALESCE(purchase_price, 0)) / COALESCE(selling_price, 0)) * 100
                    ELSE 0
                END) AS avg_margin
            FROM products
            '''
        ).fetchone()
        return {
            'total': int(row['total'] or 0),
            'active': int(row['active'] or 0),
            'inactive': int(row['inactive'] or 0),
            'out_of_stock': int(row['out_of_stock'] or 0),
            'low_stock': int(row['low_stock'] or 0),
            'inventory_value': float(row['inventory_value'] or 0),
            'avg_margin': float(row['avg_margin'] or 0),
        }
    finally:
        connection.close()


def export_products_excel(file_path: str, product_ids: Optional[List[int]] = None):
    products = list_products()
    if product_ids:
        ids = {int(pid) for pid in product_ids}
        products = [product for product in products if int(product.get('id', 0)) in ids]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'

    headers = list(_product_export_dict({}).keys())
    ws.append(headers)
    for product in products:
        row_dict = _product_export_dict(product)
        ws.append([row_dict[h] for h in headers])

    wb.save(file_path)


def export_products_csv(file_path: str, product_ids: Optional[List[int]] = None):
    products = list_products()
    if product_ids:
        ids = {int(pid) for pid in product_ids}
        products = [product for product in products if int(product.get('id', 0)) in ids]
    headers = list(_product_export_dict({}).keys())
    with open(file_path, 'w', newline='', encoding='utf-8') as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for product in products:
            writer.writerow(_product_export_dict(product))


def export_products_json(file_path: str, product_ids: Optional[List[int]] = None):
    products = list_products()
    if product_ids:
        ids = {int(pid) for pid in product_ids}
        products = [product for product in products if int(product.get('id', 0)) in ids]
    payload = [_product_export_dict(product) for product in products]
    with open(file_path, 'w', encoding='utf-8') as handle:
        json.dump(payload, handle, indent=2)
        handle.write('\n')


def get_settings() -> Dict:
    connection = get_connection()
    try:
        row = connection.execute('SELECT * FROM settings ORDER BY id DESC LIMIT 1').fetchone()
        if row:
            return dict(row)
        return {}
    finally:
        connection.close()


def update_settings(payload: Dict) -> Dict:
    current = get_settings()
    merged = {
        'github_repo': payload.get('github_repo', current.get('github_repo', 'PatelStoresOrder')),
        'branch': payload.get('branch', current.get('branch', 'main')),
        'website_url': payload.get('website_url', current.get('website_url', '')),
        'images_folder': payload.get('images_folder', current.get('images_folder', 'images')),
        'backup_folder': payload.get('backup_folder', current.get('backup_folder', 'backup')),
        'commit_message': payload.get('commit_message', current.get('commit_message', 'Update catalog')),
        'website_folder': payload.get('website_folder', current.get('website_folder', '.')),
        'database_path': payload.get('database_path', current.get('database_path', 'patelstores.db')),
        'theme': payload.get('theme', current.get('theme', 'dark')),
    }

    connection = get_connection()
    try:
        row = connection.execute('SELECT id FROM settings ORDER BY id DESC LIMIT 1').fetchone()
        if row:
            connection.execute(
                '''
                UPDATE settings
                SET github_repo = ?, branch = ?, website_url = ?, images_folder = ?, backup_folder = ?,
                    commit_message = ?, website_folder = ?, database_path = ?, theme = ?
                WHERE id = ?
                ''',
                (
                    merged['github_repo'],
                    merged['branch'],
                    merged['website_url'],
                    merged['images_folder'],
                    merged['backup_folder'],
                    merged['commit_message'],
                    merged['website_folder'],
                    merged['database_path'],
                    merged['theme'],
                    row['id'],
                ),
            )
        else:
            connection.execute(
                '''
                INSERT INTO settings (
                    github_repo, branch, website_url, images_folder, backup_folder,
                    commit_message, website_folder, database_path, theme
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (
                    merged['github_repo'],
                    merged['branch'],
                    merged['website_url'],
                    merged['images_folder'],
                    merged['backup_folder'],
                    merged['commit_message'],
                    merged['website_folder'],
                    merged['database_path'],
                    merged['theme'],
                ),
            )
        connection.commit()
    finally:
        connection.close()

    return get_settings()


def create_backup_snapshot() -> Path:
    stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    backup_path = BACKUP_DIR / f'publish-backup-{stamp}'
    backup_path.mkdir(parents=True, exist_ok=True)

    db_target = backup_path / DB_PATH.name
    shutil.copy2(DB_PATH, db_target)

    json_path = ROOT_DIR / 'products.json'
    if json_path.exists():
        shutil.copy2(json_path, backup_path / 'products.json')

    images_target = backup_path / 'images'
    if IMAGES_DIR.exists():
        for source_path in IMAGES_DIR.iterdir():
            if source_path.is_file():
                optimize_image_copy(source_path, images_target / source_path.name)

    return backup_path


def publish_changes(progress: Callable[[int, str], None]) -> Dict:
    settings = get_settings()
    branch = settings.get('branch') or 'main'
    commit_message = settings.get('commit_message') or 'Update catalog'

    def step(percent: int, message: str):
        logger.info('Publish %s%% - %s', percent, message)
        progress(percent, message)

    step(0, 'Preparing publish')
    _require_git_identity()
    _require_git_remote(branch)

    step(5, 'Saving SQLite database')
    save_sqlite_database()

    step(15, 'Creating backup')
    backup_path = create_backup_snapshot()

    step(30, 'Generating products.json')
    generate_catalog_json()

    step(45, 'Running git add')
    _git_run(['add', '.'])

    step(65, 'Running git commit')
    commit_proc = _git_run(['commit', '-m', commit_message], allow_failure=True)

    commit_output = (commit_proc.stdout + '\n' + commit_proc.stderr).lower()
    if commit_proc.returncode != 0 and 'nothing to commit' not in commit_output:
        raise PublishError(commit_proc.stderr.strip() or commit_proc.stdout.strip() or 'Git commit failed.')

    if commit_proc.returncode == 0:
        step(85, 'Running git push')
        _git_run(['push', 'origin', branch])
    else:
        step(85, 'No changes to push')

    step(100, 'Publish completed')
    return {
        'success': True,
        'backup_path': str(backup_path),
        'branch': branch,
        'commit_output': commit_proc.stdout.strip() or commit_proc.stderr.strip() or 'Nothing new to commit.',
    }
