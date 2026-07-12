import json
import os
import re
from openpyxl import load_workbook

xlsx_path = 'products.xlsx'
json_path = 'products.json'

if not os.path.exists(xlsx_path):
    raise FileNotFoundError(xlsx_path)

wb = load_workbook(xlsx_path, data_only=True)
ws = wb.active

headers = []
for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
    headers.append('' if cell is None else str(cell).strip())

header_lookup = {header.lower(): idx for idx, header in enumerate(headers) if header}

name_idx = header_lookup.get('product name')
category_idx = header_lookup.get('category')
price_idx = header_lookup.get('price')
photo_idx = header_lookup.get('photo')

if name_idx is None or category_idx is None or price_idx is None or photo_idx is None:
    raise ValueError(f'Could not map required columns. Headers found: {headers}')

products = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(cell is not None and str(cell).strip() != '' for cell in row):
        continue

    name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] is not None else ''
    category = str(row[category_idx]).strip() if category_idx < len(row) and row[category_idx] is not None else ''
    price_raw = row[price_idx] if price_idx < len(row) else None
    photo = str(row[photo_idx]).strip() if photo_idx < len(row) and row[photo_idx] is not None else ''

    if not name:
        continue

    try:
        price_value = float(price_raw)
    except (TypeError, ValueError):
        price_value = 0

    category_value = category if category else 'General'
    image_value = photo if photo else 'images/placeholder.png'
    safe_id = re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-') or 'product'

    products.append({
        'id': safe_id,
        'name': name,
        'category': category_value,
        'price': round(price_value, 2),
        'image': image_value,
        'description': ''
    })

with open(json_path, 'w', encoding='utf-8') as f:
    json.dump({'products': products}, f, indent=2)
    f.write('\n')

print(f'Wrote {len(products)} products to {json_path}')
